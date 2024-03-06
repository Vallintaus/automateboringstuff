import openpyxl

def invert_spreadsheet(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    # Determine the range of the existing data
    max_row = sheet.max_row
    max_column = sheet.max_column

    # Iterate through each cell in the original sheet
    for row in range(1, max_row + 1):
        for col in range(1, max_column + 1):
            # Read the value from old sheet
            original_value = sheet.cell(row = row, column = col).value

            # Write the value into the new sheet swapping row and col
            new_sheet.cell(row = col, column = row).value = original_value

    new_filename = 'inverted' + filename
    new_wb.save(new_filename)
    print(f"Inverted spreadsheet saved as '{new_filename}'")

if __name__ == '__main__':
    import sys
    if len(sys.argv) == 2:
        filename = sys.argv[1]
        try:
            invert_spreadsheet(filename)
        except Exception as e:
            print(f"An error occurred: {e}")
    else:
        print(f"Usage: python3 <script.py> <filename.xlsx>")