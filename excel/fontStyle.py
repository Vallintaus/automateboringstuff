# Example: Change bold every row with cost per pound > 5$

import openpyxl
from openpyxl.styles import Font   # This allows you to type Font() instead of openpyxl.styles.Font()


wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active

bold16Font = Font(size=12, bold=True)

cost_per_pound_column = 'B'

for row in range(2, sheet.max_row + 1):
    cell = sheet[f'{cost_per_pound_column}{row}']

    if cell.value and float(cell.value) > 5:
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row, column=col).font = bold16Font

wb.save("styled_produceSales.xlsx")