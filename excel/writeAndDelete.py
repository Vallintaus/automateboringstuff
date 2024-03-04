import openpyxl

# Create a workbook
wb = openpyxl.Workbook()

# Get the active worksheet
ws = wb.active

# Rename worksheet
ws.title = "ExampleTitle"

# Add data to cells
ws['A1'] = "hello"
ws['A2'] = "World"

# Adding a row of data
ws.append(['This', 'is', 'a', 'new', 'row'])

# Create new sheet
wb.create_sheet(title="NewSheet")
# Create sheet at the first position
wb.create_sheet(title="FirstSheet", index=0)

# Save the workbook to a file
wb.save('MyExampleFile.xlsx')
# This will save the file to current working directory

