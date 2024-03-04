import openpyxl
import os

#Get current working directory
current_dir = os.getcwd()

# Join the current directory with the filename
file_path = os.path.join(current_dir, 'excel', 'example.xlsx')

wb = openpyxl.load_workbook(file_path)
sheet = wb['Sheet1']  # Directly using the sheet name

cell_a1 = sheet['A1']
print(cell_a1) # Output: <Cell Sheet1.A1>
print(cell_a1.value) # Output: content of cell A1

# Access another cell
c = sheet['B1']
print(c.value)

# Formating strings with cell properties
print(f"Row {c.row}, Column: {c.column_letter} is {c.value}")
# Output: 'Row 1, Column B is Apples'

print(f"Cell {c.coordinate} is {c.value}")
# Output: 'Cell B1 is Apples'

# Accessing and printing another cell's value
print(sheet['C1'].value)

print(sheet.cell(row=1, column=2).value)
# output: Apples

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)
# output: 
# 1 Apples
# 3 Pears
# 5 Apples
# 7 Strawberries
    

# Selecting multiple rows/columns
# print(tuple(sheet['A1':'C3']))
# Output: ((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>), (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>), (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>))

for rowOfCellObjects in sheet ['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')
# Output A1 2015-04-05 13:34:02
# B1 Apples
# C1 73
# --- END OF ROW ---
# A2 2015-04-05 03:41:23
# B2 Cherries
# C2 85
# --- END OF ROW ---
# A3 2015-04-06 12:46:51
# B3 Pears
# C3 14
# --- END OF ROW ---