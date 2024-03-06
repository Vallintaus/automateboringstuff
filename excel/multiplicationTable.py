# Creates workbook that takes a number N from the command line and creates an N×N multiplication table in an Excel

import openpyxl
from openpyxl.styles import Font
import sys

def create_multiplication_table(n):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.title = "Multiplication Table"

    bold_font = Font(bold=True)

    for row in range(1, n + 1):
        for col in range(1, n + 1):
            # Set row and colum headers
            if row == 1:
                ws.cell(row = row, column= col + 1, value = col).font = bold_font
            if col == 1:
                ws.cell(row = row + 1, column = col, value = row).font = bold_font
            
            # Calculate and write the product
            ws.cell(row = row + 1, column = col + 1, value = row * col)
    
    wb.save('MultiplicationTable.xlsx')

if __name__ == "__main__":
    if len(sys.argv) == 2:
        try:
            n = int(sys.argv[1])
            create_multiplication_table(n)
            print(f"{n}x{n} Multiplication Table created successfully in 'MultiplicationTable.xlsx'")
        except ValueError:
            print("Please enter a valid integer")
    else:
        print("Usage: python multiplicationTable.py N")
