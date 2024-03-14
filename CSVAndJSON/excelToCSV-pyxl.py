import csv
from openpyxl import load_workbook
import os

def excel_to_csv(excel_file):
    workbook = load_workbook(excel_file)

    # create output dir
    output_dir = "csv_output"
    os.makedirs(output_dir, exist_ok=True)

    for sheet_name in workbook.sheetnames:
        # Load worksheet
        worksheet = workbook[sheet_name]

        # Name CSV file after excel file
        csv_file_name = f"{output_dir}/{sheet_name}.csv"

        # Open CSV file and write content of worksheet
        with open(csv_file_name, 'w', newline='', encoding='utf-8') as csv_file:
            writer = csv.writer(csv_file)

            for row in worksheet.iter_rows(values_only=True):
                writer.writerow(row)

        print(f"Worksheet '{sheet_name}' has been converted to '{csv_file_name}'")

# Replace 'your_excel_file.xlsx' with the path to your Excel file
excel_file = 'your_excel_file.xlsx'
excel_to_csv(excel_file)
